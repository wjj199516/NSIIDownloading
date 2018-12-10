from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.request import urlretrieve
from urllib.error import HTTPError
from bs4 import BeautifulSoup
import time
import os
from openpyxl import load_workbook
from openpyxl import Workbook

def getSingleUrl(driver,url,name):
    try:
        driver.get(url)
        time.sleep(0.5)
        #WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "specimen_info")))
        pageSource = driver.page_source
        # print(pageSource)
        bsObj = BeautifulSoup(pageSource, 'html.parser')

        try:
            science_name = bsObj.find('div', {"id": "specimen_info"}).find('a', {"id": "sp_link"}).find('span', {"class": "lname"}).get_text().strip()
        except AttributeError:
            print('No science_name')
            science_name = 'None'
        #print('science_name', science_name)
        try:
            chinese_name = bsObj.find('div', {"id": "specimen_info"}).find('a', {"id": "sp_link"}).find('span', {
                "class": "cname"}).get_text().strip()
        except AttributeError:
            print('No chinese_name')
            chinese_name = 'None'
        #print("chinese_name", chinese_name)
        try:
            collector = bsObj.find('div',{"id":"specimen_info"}).find('div',{'class':'collect_info'}).findAll('span',{"class":"field_value"})[0].get_text().strip()
        except AttributeError:
            print('No collector')
            collector = 'None'
        #print("collector", collector)
        try:
            collector_num = bsObj.find('div', {"id": "specimen_info"}).find('div', {'class': 'collect_info'}).findAll('span', {"class": "field_value"})[1].get_text().strip()
        except AttributeError:
            print('No collector_num')
            collector_num = 'None'
    #print("collector_num", collector_num)
        try:
            collector_date = bsObj.find('div', {"id": "specimen_info"}).find('div', {'class': 'collect_info'}).findAll('span', {"class": "field_value"})[2].get_text().strip()
        except AttributeError:
            print('No collector_date')
            collector_date = 'None'
    #print("collector_date", collector_date)
        try:
            collector_loc = bsObj.find('div', {"id": "specimen_info"}).find('div', {'class': 'collect_info'}).findAll('span', {"class": "field_value"})[3].get_text().strip()
        except AttributeError:
            print('No collector_loc')
            collector_loc = 'None'
    #print("collector_loc", collector_loc)
        try:
            GUID = bsObj.find('div', {"id": "specimen_info"}).find('div', {'class': 'meta'}).findAll('span', {"class": "field_value"})[0].get_text().strip()
        except AttributeError:
            print('No GUID')
            GUID = 'None'
    #print("GUID", GUID)
        try:
            herbarium = bsObj.find('div', {"id": "specimen_info"}).find('div', {'class': 'meta'}).findAll('span', {"class": "field_value"})[1].get_text().strip()
        except AttributeError:
            print('No herbarium')
            herbarium = 'None'
    #print("herbarium", herbarium)
        try:
            provider = bsObj.find('div', {"id": "specimen_info"}).find('div', {'class': 'meta'}).findAll('span', {"class": "field_value"})[2].get_text().strip()
        except AttributeError:
            print('No provider')
            provider = 'None'
    #print("provider", provider)
        try:
            image_url = bsObj.find("img", {"class": "box-shadow-1"})["src"]
            #print(image_url)
            #print(image_url)
                #        print("image_url", image_url)
            #image_path = getDownloadPath(baseUrl, image_url, downloadDirectory)
            #print(image_path)
            #path = image_url.replace(baseUrl, "")
            path = image_url.split('/')[-1]
            #print(path)
            path = downloadDirectory + path
            directory = os.path.dirname(path)
            #print(directory)
            if not os.path.exists(directory):
                os.makedirs(directory)
            if image_url != '':
                #image_path = '/Users/Simon/Documents/DATADIARY/20180822/Acer_records/AcerImage/'
                try:
                    urlretrieve(image_url, path)
                except HTTPError:
                    image_url = 'None'
                    image_path = 'None'
        except TypeError as e:
            image_url = 'None'
            image_path = 'None'

        duplicated_list = [science_name,chinese_name,collector,collector_num,collector_date,collector_loc,GUID,herbarium,provider]
        None_list = ['None','None']
        if set(duplicated_list) != set(None_list) :
            #file = open('D:\\ZGFSft\\Environment\\NSII\\GetResults\\' + name ,'a')
            #file.write(science_name + '\t' + chinese_name + '\t' + collector + '\t' + collector_num + '\t' + collector_date + '\t' + collector_loc + '\t' + GUID + '\t' + herbarium + '\t' + provider + '\t' + url + '\n')
             #print(science_name + '\t' + chinese_name + '\t' + collector + '\t' + collector_num + '\t' + collector_date + '\t' + collector_loc + '\t' + GUID + '\t' + herbarium + '\t' + provider + '\t' + url + '\n')
            #file.close()
            wb = load_workbook('D:\\ZGFSft\\Environment\\NSII\\GetResults\\' + name + '.xlsx')
            sheet = wb['Sheet']
            row = [science_name, chinese_name, collector, collector_num, collector_date, collector_loc, GUID, herbarium,
                   provider, url]
            # print(row)
            sheet.append(row)
            rownumber = sheet.max_row
            sheet['K' + str(rownumber)].hyperlink = path.replace('D:\\ZGFSft\\Environment\\NSII\\GetResults\\', '')
            wb.save('D:\\ZGFSft\\Environment\\NSII\\GetResults\\' + name + '.xlsx')
            gottenfile.write(line)
    except IndexError:
        getSingleUrl(driver, url, name)

#物种链接名单
baseUrl = 'http://img.cvh.ac.cn/imgcvh/'
downloadDirectory = 'D:\\ZGFSft\\Environment\\NSII\\GetResults\\Image\\'


name = 	'Polystichum glaciale_2018-08-30_5'
driver = webdriver.PhantomJS(executable_path='D:\\ZGFSft\\Environment\\Python\\Lib\\site-packages\\phantomjs-2.1.1-windows\\bin\\phantomjs.exe')
infile = open('D:\\ZGFSft\\Environment\\NSII\links\\' + name,'rU')
readgottenfile = open('D:\\ZGFSft\\Environment\\NSII\links\\GottenLinks.txt','rU')
gottenCotent = readgottenfile.read()
readgottenfile.close()
gottenfile = open('D:\\ZGFSft\\Environment\\NSII\links\\GottenLinks.txt','a')

while True:
    line = infile.readline()
    if line:
        if line not in gottenCotent:
            print(line)
            url = line[:-1]
            getSingleUrl(driver,url,name)
    else:
        break
infile.close()
gottenfile.close()
driver.quit()

